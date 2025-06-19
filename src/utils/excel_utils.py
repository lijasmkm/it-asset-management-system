"""
Excel Utilities for IT Asset Management System
Handles import and export of data to/from Excel
"""
import os
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from datetime import datetime
from tkinter import filedialog, messagebox
import tkinter as tk

class ExcelUtils:
    def __init__(self, asset_controller=None):
        """
        Initialize Excel utilities
        
        Args:
            asset_controller: Asset controller instance
        """
        self.asset_controller = asset_controller
        
        # Create exports and templates directories if they don't exist
        self.exports_dir = os.path.join(os.path.dirname(os.path.dirname(os.path.dirname(__file__))), "exports")
        self.templates_dir = os.path.join(os.path.dirname(os.path.dirname(os.path.dirname(__file__))), "templates")
        
        os.makedirs(self.exports_dir, exist_ok=True)
        os.makedirs(self.templates_dir, exist_ok=True)
    
    def export_assets_to_excel(self, assets, filename=None):
        """
        Export asset data to Excel
        
        Args:
            assets (list): List of asset dictionaries
            filename (str, optional): Output filename
        
        Returns:
            str: Path to the exported file
        """
        # Generate filename if not provided
        if not filename:
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            filename = f"asset_export_{timestamp}.xlsx"
        
        # Create a new workbook and select the active sheet
        wb = openpyxl.Workbook()
        sheet = wb.active
        sheet.title = "Assets"
        
        # Define headers based on asset fields
        headers = [
            "Serial Number", "Company", "Location", "Category", "Status",
            "Username", "Designation", "Department", "Model", "Description",
            "Issue Date", "Computer ID", "Working Status", "Condition",
            "Audit", "Employee ID", "Purchase Date", "Rack/Tray Number",
            "Service Center", "LPO Number", "Invoice Number", "Supplier",
            "Estimated Cost", "Remarks"
        ]
        
        # Add headers with formatting
        for col_num, header in enumerate(headers, 1):
            cell = sheet.cell(row=1, column=col_num)
            cell.value = header
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        
        # Add data rows
        for row_num, asset in enumerate(assets, 2):
            sheet.cell(row=row_num, column=1).value = asset.get("serial_number", "")
            sheet.cell(row=row_num, column=2).value = asset.get("company", "")
            sheet.cell(row=row_num, column=3).value = asset.get("location", "")
            sheet.cell(row=row_num, column=4).value = asset.get("category", "")
            sheet.cell(row=row_num, column=5).value = asset.get("status", "")
            sheet.cell(row=row_num, column=6).value = asset.get("username", "")
            sheet.cell(row=row_num, column=7).value = asset.get("designation", "")
            sheet.cell(row=row_num, column=8).value = asset.get("department", "")
            sheet.cell(row=row_num, column=9).value = asset.get("model", "")
            sheet.cell(row=row_num, column=10).value = asset.get("description", "")
            sheet.cell(row=row_num, column=11).value = asset.get("issue_date", "")
            sheet.cell(row=row_num, column=12).value = asset.get("computer_id", "")
            sheet.cell(row=row_num, column=13).value = asset.get("working_status", "")
            sheet.cell(row=row_num, column=14).value = asset.get("condition", "")
            sheet.cell(row=row_num, column=15).value = asset.get("audit", "")
            sheet.cell(row=row_num, column=16).value = asset.get("employee_id", "")
            sheet.cell(row=row_num, column=17).value = asset.get("purchase_date", "")
            sheet.cell(row=row_num, column=18).value = asset.get("rack_tray_number", "")
            sheet.cell(row=row_num, column=19).value = asset.get("service_center", "")
            sheet.cell(row=row_num, column=20).value = asset.get("lpo_number", "")
            sheet.cell(row=row_num, column=21).value = asset.get("invoice_number", "")
            sheet.cell(row=row_num, column=22).value = asset.get("supplier", "")
            sheet.cell(row=row_num, column=23).value = asset.get("estimated_cost", "")
            sheet.cell(row=row_num, column=24).value = asset.get("remarks", "")
        
        # Auto-adjust column widths
        for column in sheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            adjusted_width = max_length + 2
            sheet.column_dimensions[column_letter].width = adjusted_width
        
        # Save the workbook
        output_path = os.path.join(self.exports_dir, filename)
        wb.save(output_path)
        
        return output_path
    
    def import_assets_from_excel(self, file_path=None):
        """
        Import assets from Excel file
        
        Args:
            file_path (str, optional): Path to Excel file
        
        Returns:
            tuple: (success, message, imported_count)
        """
        if not self.asset_controller:
            return False, "Asset controller not initialized", 0
        
        # If file path not provided, open file dialog
        if not file_path:
            file_path = filedialog.askopenfilename(
                title="Select Excel File",
                filetypes=[("Excel files", "*.xlsx *.xls")],
                initialdir=os.path.expanduser("~")
            )
            
            if not file_path:  # User cancelled
                return False, "Import cancelled", 0
        
        try:
            # Load the workbook
            wb = openpyxl.load_workbook(file_path)
            sheet = wb.active
            
            # Get headers from the first row
            headers = [cell.value for cell in sheet[1]]
            
            # Map Excel headers to database fields
            field_mapping = {
                "Serial Number": "serial_number",
                "Company": "company",
                "Location": "location",
                "Category": "category",
                "Status": "status",
                "Username": "username",
                "Designation": "designation",
                "Department": "department",
                "Model": "model",
                "Description": "description",
                "Issue Date": "issue_date",
                "Computer ID": "computer_id",
                "Working Status": "working_status",
                "Condition": "condition",
                "Audit": "audit",
                "Employee ID": "employee_id",
                "Purchase Date": "purchase_date",
                "Rack/Tray Number": "rack_tray_number",
                "Service Center": "service_center",
                "LPO Number": "lpo_number",
                "Invoice Number": "invoice_number",
                "Supplier": "supplier",
                "Estimated Cost": "estimated_cost",
                "Remarks": "remarks"
            }
            
            # Validate required headers
            required_fields = ["Serial Number", "Category"]
            for field in required_fields:
                if field not in headers:
                    return False, f"Required field '{field}' not found in Excel file", 0
            
            # Process data rows
            imported_count = 0
            errors = []
            
            for row_num, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), 2):
                # Skip empty rows
                if all(cell is None or cell == "" for cell in row):
                    continue
                
                # Create asset data dictionary
                asset_data = {}
                for col_num, value in enumerate(row):
                    if col_num < len(headers) and headers[col_num] in field_mapping:
                        db_field = field_mapping[headers[col_num]]
                        asset_data[db_field] = value
                
                # Validate required fields
                if not asset_data.get("serial_number"):
                    errors.append(f"Row {row_num}: Missing Serial Number")
                    continue
                    
                if not asset_data.get("category"):
                    errors.append(f"Row {row_num}: Missing Category")
                    continue
                
                # Check for duplicate serial numbers
                existing_asset = self.asset_controller.get_asset_by_serial(asset_data["serial_number"])
                if existing_asset:
                    errors.append(f"Row {row_num}: Duplicate Serial Number '{asset_data['serial_number']}'")
                    continue
                
                # Add the asset
                success, message = self.asset_controller.add_asset(asset_data)
                if success:
                    imported_count += 1
                else:
                    errors.append(f"Row {row_num}: {message}")
            
            # Return results
            if errors:
                return True, f"Imported {imported_count} assets with {len(errors)} errors:\n" + "\n".join(errors), imported_count
            else:
                return True, f"Successfully imported {imported_count} assets", imported_count
                
        except Exception as e:
            return False, f"Import error: {str(e)}", 0
    
    def create_import_template(self):
        """
        Create an import template file
        
        Returns:
            str: Path to the template file
        """
        # Create a new workbook and select the active sheet
        wb = openpyxl.Workbook()
        sheet = wb.active
        sheet.title = "Asset Import Template"
        
        # Define headers based on asset fields
        headers = [
            "Serial Number", "Company", "Location", "Category", "Status",
            "Username", "Designation", "Department", "Model", "Description",
            "Issue Date", "Computer ID", "Working Status", "Condition",
            "Audit", "Employee ID", "Purchase Date", "Rack/Tray Number",
            "Service Center", "LPO Number", "Invoice Number", "Supplier",
            "Estimated Cost", "Remarks"
        ]
        
        # Add headers with formatting
        for col_num, header in enumerate(headers, 1):
            cell = sheet.cell(row=1, column=col_num)
            cell.value = header
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
            
            # Mark required fields
            if header in ["Serial Number", "Category"]:
                cell.fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
        
        # Add instructions row
        sheet.insert_rows(1)
        instruction_cell = sheet.cell(row=1, column=1)
        instruction_cell.value = "Instructions: Fill in the asset details below. Fields marked in red are required."
        instruction_cell.font = Font(bold=True)
        sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
        
        # Add example data
        example_data = [
            "SN12345678", "Meraki", "SS7", "Laptop", "Stock",
            "", "", "", "Dell Latitude 5420", "Core i5, 8GB RAM, 256GB SSD",
            "", "", "Working", "New",
            "", "", "2025-01-15", "",
            "", "LPO-2025-001", "INV-2025-001", "Dell",
            "1200", ""
        ]
        
        for col_num, value in enumerate(example_data, 1):
            sheet.cell(row=3, column=col_num).value = value
        
        # Auto-adjust column widths
        for column in sheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            adjusted_width = max_length + 2
            sheet.column_dimensions[column_letter].width = adjusted_width
        
        # Save the workbook
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f"asset_import_template_{timestamp}.xlsx"
        output_path = os.path.join(self.templates_dir, filename)
        wb.save(output_path)
        
        return output_path
    
    def create_asset_issue_form(self, asset=None):
        """
        Create an IT Asset Hardware Issue Form
        
        Args:
            asset (dict, optional): Asset data to pre-fill
        
        Returns:
            str: Path to the form file
        """
        # Create a new workbook and select the active sheet
        wb = openpyxl.Workbook()
        sheet = wb.active
        sheet.title = "IT Asset Issue Form"
        
        # Add title
        sheet.merge_cells('A1:H1')
        title_cell = sheet.cell(row=1, column=1)
        title_cell.value = "IT ASSET HARDWARE ISSUE FORM"
        title_cell.font = Font(bold=True, size=14)
        title_cell.alignment = Alignment(horizontal="center")
        
        # Add form fields
        form_fields = [
            ("Date:", ""),
            ("Employee Name:", asset.get("username", "") if asset else ""),
            ("Employee ID:", asset.get("employee_id", "") if asset else ""),
            ("Department:", asset.get("department", "") if asset else ""),
            ("Designation:", asset.get("designation", "") if asset else ""),
            ("", ""),
            ("Asset Details", ""),
            ("Category:", asset.get("category", "") if asset else ""),
            ("Serial Number:", asset.get("serial_number", "") if asset else ""),
            ("Model:", asset.get("model", "") if asset else ""),
            ("Description:", asset.get("description", "") if asset else ""),
            ("Computer ID:", asset.get("computer_id", "") if asset else ""),
            ("Condition:", asset.get("condition", "") if asset else ""),
            ("", ""),
            ("Acknowledgement", ""),
            ("I acknowledge receipt of the above IT asset and agree to use it for official purposes only.", ""),
            ("", ""),
            ("Employee Signature:", ""),
            ("IT Department Signature:", "")
        ]
        
        for row_num, (label, value) in enumerate(form_fields, 3):
            if label:
                label_cell = sheet.cell(row=row_num, column=1)
                label_cell.value = label
                if "Details" in label or "Acknowledgement" in label:
                    label_cell.font = Font(bold=True)
                    sheet.merge_cells(f'A{row_num}:H{row_num}')
                    label_cell.alignment = Alignment(horizontal="center")
                else:
                    label_cell.font = Font(bold=True)
                    value_cell = sheet.cell(row=row_num, column=3)
                    value_cell.value = value
                    sheet.merge_cells(f'C{row_num}:H{row_num}')
        
        # Auto-adjust column widths
        sheet.column_dimensions['A'].width = 20
        sheet.column_dimensions['B'].width = 5
        sheet.column_dimensions['C'].width = 30
        
        # Save the workbook
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f"asset_issue_form_{timestamp}.xlsx"
        if asset and asset.get("serial_number"):
            filename = f"asset_issue_form_{asset.get('serial_number')}_{timestamp}.xlsx"
        
        output_path = os.path.join(self.exports_dir, filename)
        wb.save(output_path)
        
        return output_path
    
    def create_asset_transfer_form(self, asset=None):
        """
        Create an IT Asset Transfer Form
        
        Args:
            asset (dict, optional): Asset data to pre-fill
        
        Returns:
            str: Path to the form file
        """
        # Create a new workbook and select the active sheet
        wb = openpyxl.Workbook()
        sheet = wb.active
        sheet.title = "IT Asset Transfer Form"
        
        # Add title
        sheet.merge_cells('A1:H1')
        title_cell = sheet.cell(row=1, column=1)
        title_cell.value = "IT ASSET TRANSFER FORM"
        title_cell.font = Font(bold=True, size=14)
        title_cell.alignment = Alignment(horizontal="center")
        
        # Add form fields
        form_fields = [
            ("Date:", ""),
            ("", ""),
            ("Asset Details", ""),
            ("Category:", asset.get("category", "") if asset else ""),
            ("Serial Number:", asset.get("serial_number", "") if asset else ""),
            ("Model:", asset.get("model", "") if asset else ""),
            ("Description:", asset.get("description", "") if asset else ""),
            ("Computer ID:", asset.get("computer_id", "") if asset else ""),
            ("", ""),
            ("From", ""),
            ("Employee Name:", asset.get("username", "") if asset else ""),
            ("Employee ID:", asset.get("employee_id", "") if asset else ""),
            ("Department:", asset.get("department", "") if asset else ""),
            ("", ""),
            ("To", ""),
            ("Employee Name:", ""),
            ("Employee ID:", ""),
            ("Department:", ""),
            ("", ""),
            ("Reason for Transfer:", ""),
            ("", ""),
            ("Signatures", ""),
            ("Current User:", ""),
            ("New User:", ""),
            ("IT Department:", "")
        ]
        
        for row_num, (label, value) in enumerate(form_fields, 3):
            if label:
                label_cell = sheet.cell(row=row_num, column=1)
                label_cell.value = label
                if label in ["Asset Details", "From", "To", "Signatures"]:
                    label_cell.font = Font(bold=True)
                    sheet.merge_cells(f'A{row_num}:H{row_num}')
                    label_cell.alignment = Alignment(horizontal="center")
                else:
                    label_cell.font = Font(bold=True)
                    value_cell = sheet.cell(row=row_num, column=3)
                    value_cell.value = value
                    sheet.merge_cells(f'C{row_num}:H{row_num}')
        
        # Auto-adjust column widths
        sheet.column_dimensions['A'].width = 20
        sheet.column_dimensions['B'].width = 5
        sheet.column_dimensions['C'].width = 30
        
        # Save the workbook
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f"asset_transfer_form_{timestamp}.xlsx"
        if asset and asset.get("serial_number"):
            filename = f"asset_transfer_form_{asset.get('serial_number')}_{timestamp}.xlsx"
        
        output_path = os.path.join(self.exports_dir, filename)
        wb.save(output_path)
        
        return output_path
    
    def open_file(self, file_path):
        """
        Open a file with the default application
        
        Args:
            file_path (str): Path to the file
        """
        try:
            import os
            os.startfile(file_path)
            return True
        except Exception as e:
            print(f"Error opening file: {e}")
            return False
